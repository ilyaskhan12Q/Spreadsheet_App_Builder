"""
renderers/xlsx_writer.py
Primary renderer using openpyxl to generate .xlsx workbooks.
"""
import io
import logging

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from core.blueprint import Blueprint
from renderers.style_mapper import get_openpyxl_styles

logger = logging.getLogger("sab.xlsx_writer")


class XlsxRenderer:
    def __init__(self) -> None:
        self.wb = openpyxl.Workbook()
        self.ws: Worksheet = self.wb.active  # type: ignore

    def render(
        self, blueprint: Blueprint, save_path: str | None = None
    ) -> bytes | None:
        """
        Renders the blueprint to an openpyxl workbook.
        If save_path is provided, saves to that path.
        Otherwise, returns the bytes of the .xlsx file.
        """
        self.ws.title = blueprint.meta.title[:31]  # Excel limits sheet name to 31 chars
        self.ws.sheet_view.showGridLines = not blueprint.meta.hide_gridlines

        # Set column widths
        if blueprint.meta.col_widths:
            for col_letter, width in blueprint.meta.col_widths.items():
                self.ws.column_dimensions[col_letter].width = width

        # Set row heights
        if blueprint.meta.row_heights:
            for row_idx_str, height in blueprint.meta.row_heights.items():
                try:
                    row_idx = int(row_idx_str)
                    self.ws.row_dimensions[row_idx].height = height
                except ValueError:
                    pass

        # Write cells
        for cell in blueprint.cells:
            xl_cell = self.ws[cell.cell_id]

            # Value or formula
            if cell.formula:
                xl_cell.value = cell.formula
            elif cell.value is not None:
                xl_cell.value = cell.value

            # Styling
            if cell.style:
                styles = get_openpyxl_styles(cell.style)
                if 'font' in styles:
                    xl_cell.font = styles['font']
                if 'alignment' in styles:
                    xl_cell.alignment = styles['alignment']
                if 'fill' in styles:
                    xl_cell.fill = styles['fill']
                if 'border' in styles:
                    xl_cell.border = styles['border']

                # Number format
                if cell.style.number_format:
                    xl_cell.number_format = cell.style.number_format

            # Validation
            if cell.validation:
                if cell.validation.type == "list" and cell.validation.formula1:
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{cell.validation.formula1}"',
                        allow_blank=True
                    )
                    self.ws.add_data_validation(dv)
                    dv.add(xl_cell)

        # Merges
        if blueprint.merges:
            for merge in blueprint.merges:
                self.ws.merge_cells(merge.range)

        # Save or return bytes
        if save_path:
            self.wb.save(save_path)
            logger.info(f"Saved xlsx to {save_path}")
            return None
        else:
            buffer = io.BytesIO()
            self.wb.save(buffer)
            return buffer.getvalue()
