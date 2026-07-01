import json
from dataclasses import dataclass, field
from typing import Any


@dataclass
class SpreadsheetContext:
    used_range: str = "A1"
    headers: list[str] = field(default_factory=list)
    data_sample: list[dict[str, Any]] = field(default_factory=list)
    named_ranges: dict[str, str] = field(default_factory=dict)
    existing_styles: dict[str, Any] = field(default_factory=dict)
    sheet_names: list[str] = field(default_factory=list)

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "SpreadsheetContext":
        """
        Build a SpreadsheetContext from a dictionary, ensuring type alignment.
        """
        return cls(
            used_range=data.get("used_range", "A1"),
            headers=data.get("headers", []),
            data_sample=data.get("data_sample", []),
            named_ranges=data.get("named_ranges", {}),
            existing_styles=data.get("existing_styles", {}),
            sheet_names=data.get("sheet_names", [])
        )

    def to_prompt_string(self) -> str:
        """
        Serializes context information for embedding into the LLM system/user prompts.
        """
        lines = [
            "### SPREADSHEET CONTEXT INFO ###",
            f"Active sheet used range: {self.used_range}",
            f"Sheet names: {', '.join(self.sheet_names) if self.sheet_names else 'None'}",
            "Headers detected:",
            f"  {json.dumps(self.headers)}",
            "Data sample preview:",
            json.dumps(self.data_sample, indent=2),
            "Named Ranges:",
            json.dumps(self.named_ranges, indent=2),
            "Existing Styles Summary:",
            json.dumps(self.existing_styles, indent=2),
            "###############################"
        ]
        return "\n".join(lines)


class ContextScanner:
    """
    Scans the active spreadsheet state (LibreOffice or workbook payload)
    and constructs a SpreadsheetContext for AI provider prompt injection.
    """
    def build_context(self, spreadsheet_handle: Any = None) -> SpreadsheetContext:
        if not spreadsheet_handle:
            return SpreadsheetContext()

        # 1. Check if spreadsheet_handle is a file path (string or Path object)
        from pathlib import Path

        import openpyxl

        path_str = str(spreadsheet_handle)
        if path_str.endswith(".xlsx"):
            path = Path(path_str)
            if path.exists():
                try:
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    sheet_names = wb.sheetnames
                    active_sheet = wb.active

                    max_row = active_sheet.max_row or 1
                    max_col = active_sheet.max_column or 1

                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(max_col)
                    used_range = f"A1:{col_letter}{max_row}"

                    # Read headers (first row)
                    headers = []
                    first_row = next(active_sheet.iter_rows(max_row=1, values_only=True), None)
                    if first_row:
                        headers = [str(val) for val in first_row if val is not None]

                    # Read data sample (rows 2 to 5)
                    data_sample = []
                    for row in active_sheet.iter_rows(min_row=2, max_row=5, values_only=True):
                        if any(val is not None for val in row):
                            row_dict = {}
                            for i, val in enumerate(row):
                                if i < len(headers):
                                    row_dict[headers[i]] = val
                                else:
                                    row_dict[f"Column_{i+1}"] = val
                            data_sample.append(row_dict)

                    return SpreadsheetContext(
                        used_range=used_range,
                        headers=headers,
                        data_sample=data_sample,
                        sheet_names=sheet_names
                    )
                except Exception:
                    pass

        # 2. Check if spreadsheet_handle is a UNO document object (has getSheets)
        if hasattr(spreadsheet_handle, "getSheets"):
            try:
                doc = spreadsheet_handle
                sheets = doc.getSheets()
                sheet_names = list(sheets.getElementNames())

                try:
                    sheet = doc.getCurrentController().getActiveSheet()
                except Exception:
                    sheet = sheets.getByIndex(0)

                cursor = sheet.createCursor()
                cursor.gotoStartOfUsedArea(False)
                cursor.gotoEndOfUsedArea(True)

                abs_name = cursor.AbsoluteName
                used_range = "A1"
                if "." in abs_name:
                    used_range = abs_name.split(".")[-1].replace("$", "")

                headers = []
                data_sample = []

                addr = cursor.getRangeAddress()
                end_col = min(addr.EndColumn, 26)  # limit to first 26 columns
                end_row = min(addr.EndRow, 5)     # limit to first 5 rows

                # Get headers (row 0)
                for col_idx in range(end_col + 1):
                    cell = sheet.getCellByPosition(col_idx, 0)
                    val = cell.getString()
                    if not val:
                        # Fallback to float value
                        val = cell.getValue()
                        if val == 0.0:
                            val = ""
                    if val:
                        headers.append(str(val))
                    else:
                        headers.append(f"Column_{col_idx+1}")

                # Get data sample (rows 1 to end_row)
                for row_idx in range(1, end_row + 1):
                    row_dict = {}
                    row_has_data = False
                    for col_idx in range(len(headers)):
                        cell = sheet.getCellByPosition(col_idx, row_idx)
                        val = cell.getString()
                        if not val:
                            val = cell.getValue()
                            if val == 0.0:
                                val = ""
                        if val is not None and val != "":
                            row_has_data = True
                        row_dict[headers[col_idx]] = val
                    if row_has_data:
                        data_sample.append(row_dict)

                return SpreadsheetContext(
                    used_range=used_range,
                    headers=headers,
                    data_sample=data_sample,
                    sheet_names=sheet_names
                )
            except Exception:
                pass

        return SpreadsheetContext()

