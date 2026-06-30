"""Tests for renderers/xlsx_writer.py using openpyxl read-back assertions."""
import os
import tempfile

import openpyxl
import pytest

from core.validator.schema import BlueprintValidator
from renderers.xlsx_writer import XlsxRenderer


@pytest.fixture
def pos_blueprint():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    fixture_path = os.path.join(current_dir, "fixtures", "pos_blueprint.json")
    with open(fixture_path) as f:
        raw_json = f.read()
    validator = BlueprintValidator()
    return validator.validate(raw_json)


def test_xlsx_renderer_buffer_output(pos_blueprint):
    renderer = XlsxRenderer()

    buffer_bytes = renderer.render(pos_blueprint)
    assert buffer_bytes is not None
    assert isinstance(buffer_bytes, bytes)
    assert len(buffer_bytes) > 0

    # Write to a temp file to read back with openpyxl
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(buffer_bytes)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active

        # Test 1: Worksheet title
        # pos_blueprint meta.title is "Point of Sale (POS) Terminal"
        assert ws.title == "Point of Sale (POS) Terminal"

        # Test 2: Gridlines
        assert ws.sheet_view.showGridLines is True  # hide_gridlines = false in pos_blueprint

        # Test 3: Col Widths
        # col_widths: "A": 15.0, "B": 20.0
        assert ws.column_dimensions["A"].width == 15.0
        assert ws.column_dimensions["B"].width == 20.0

        # Test 4: Values and Formulas
        assert ws["A1"].value == "POINT OF SALE CHECKOUT"
        assert ws["B10"].value == "=B6*B7"
        assert ws["B6"].value == 2

        # Test 5: Merges
        # openpyxl merged_cells properties
        merges = [str(r) for r in ws.merged_cells.ranges]
        assert "A1:E2" in merges
        assert "D5:E5" in merges

        # Test 6: Styling
        # Cell A1 styling: bg_color: "#1A237E" (indigo), fg_color: "#FFFFFF", 
        # font_size: 16.0, bold: True, h_align: center
        cell_a1 = ws["A1"]
        # color values could be in 'FFRRGGBB'
        assert cell_a1.font.color.rgb == "FFFFFFFF"
        assert cell_a1.font.bold is True
        assert cell_a1.font.size == 16.0
        assert cell_a1.alignment.horizontal == "center"
        # openpyxl fills have an rgb attribute on start_color
        assert cell_a1.fill.start_color.rgb == "FF1A237E"

        # Test 7: Validation
        # Cell B5 validation: list validation with formula1="Coffee,Tea,Muffin,Croissant"
        validations = ws.data_validations.dataValidation
        assert len(validations) > 0
        dv = validations[0]
        assert dv.type == "list"
        assert dv.formula1 == '"Coffee,Tea,Muffin,Croissant"'
        assert "B5" in dv.cells

    finally:
        os.remove(tmp_path)
